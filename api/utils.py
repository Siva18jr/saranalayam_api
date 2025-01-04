from django.http import HttpResponse
from django.shortcuts import render
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status
from django.contrib.auth.models import User
from .serializers import SignUpSerializer, UserSerializer, ActivitySerializer, ProjectSerializer, ActivityImagesSerializer, WorkSerializer, AmountSerializer, FoodSerializer, DonationSerializer
from .models import AppUsers, Posts, Projects, Work, Amount, Food, Donation
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.tokens import UntypedToken
from datetime import datetime, timedelta
from django.utils import timezone
from collections import defaultdict
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.db.models import Sum


class SignUp(APIView):

    def post(self, request):

        data = {
            'username' : request.data['username'],
            'password' : request.data['password']
        }

        serializer = SignUpSerializer(data = data)
        
        if serializer.is_valid():

            serializer.save()

            user = User.objects.filter(username = request.data['username']).first()
            user.set_password(request.data['password'])
            user.save()

            token = RefreshToken.for_user(user)

            userData = {
                'username' : request.data['username'],
                'number' : request.data['number'],
                'password' : request.data['password'],
                'token' : str(token),
                'type' : request.data['type'],
                'created_by' : request.data['created_by']
            }

            modelSerializer = UserSerializer(data = userData)

            if modelSerializer.is_valid():

                modelSerializer.save()
           
                return Response({
                    'status' : True,
                    'data' : modelSerializer.data,
                    'message' : 'Account registered'
                })
            else:
           
                field_names = []

                print(modelSerializer.errors)
           
                for field_name, field_errors in modelSerializer.errors.items():
                    field_names.append(field_name)
           
                return Response({
                    'status' : False,
                    'data' : { },
                    'message' : f'Invalid data in {field_names}'
                })
        else:
           
            field_names = []

            print(serializer.errors)
    
            for field_name, field_errors in serializer.errors.items():
                field_names.append(field_name)
           
            return Response({
                'status' : False,
                'data' : { },
                'message' : f'Error : {field_names}'
            })
        

def login(request):

    username = request.query_params.get('username')

    if AppUsers.objects.filter(username=username).exists() is True:
       
        users = AppUsers.objects.get(username=username)
     
        serializer = UserSerializer(instance = users, many= False)
       
        if(serializer.data['password'] == request.query_params.get('password')):
            return Response({
                'status' : True,
                'data' : serializer.data,
                'message' : 'User Verified'
            })
        else:
            return Response({
                'status' : False,
                'data' : { },
                'message' : 'Wrong crendential'
            })

    else:
        return Response({
            'status' : False,
            'data' : { },
            'message' : 'Email not exists'
        })
    

def guestUser(request):

    posts = Posts.objects.all()
    serializer = ActivitySerializer(posts, many=True)

    return Response({
        'data' : serializer.data,
        'status' : True,
        'message' : 'Posts Fetched'
    })


def getPosts(request):

    posts = Posts.objects.all()
    serializer = ActivitySerializer(posts, many=True)

    return Response({
        'data' : serializer.data,
        'status' : True,
        'message' : 'Posts Fetched'
    })


def checkUserName(request):

    username = request.query_params.get('username')

    if AppUsers.objects.filter(username=username).exists() is True:
        return Response({
            'data' : True,
            'status' : status.HTTP_201_CREATED,
            'message' : 'Username Exists'
        })
    else:
        return Response({
            'data' : False,
            'status' : status.HTTP_201_CREATED,
            'message' : 'Username not taken'
        })
    

def createProject(request):
    
    serializer = ProjectSerializer(data=request.data)

    name = request.data['name']

    if Projects.objects.filter(name=name).exists() is True:
        return Response({
            'data' : {},
            'status' : False,
            'message' : 'Project Exists'
        })
    else:
        if serializer.is_valid():
            serializer.save()
            return Response({
                'status' : True,
                'data' : serializer.data,
                'message' : 'New Project Created'
            })
        else:
            return Response({
                'status' : False,
                'data' : serializer.data,
                'message' : 'Project not created'
            })
        

def getProjects(request):
    
    projects = Projects.objects.all()
    serializer = ProjectSerializer(projects, many=True)

    return Response({
        'data' : serializer.data,
        'status' : True,
        'message' : 'Fetched Projects'
    })


def uploadImage(request):

    serializer = ActivityImagesSerializer(data=request.data)

    if serializer.is_valid():
            serializer.save()
            return Response({
                'status' : True,
                'data' : serializer.data,
                'message' : 'Image Uploaded'
            })
    else:
        return Response({
            'status' : False,
            'data' : serializer.data,
            'message' : 'Image not Uploaded'
        })


def addActivity(request):
    
    serializer = ActivitySerializer(data=request.data)

    if serializer.is_valid():
        serializer.save()
        return Response({
            'status' : True,
            'data' : serializer.data,
            'message' : 'New Activity added'
        })
    else:
        return Response({
            'status' : False,
            'data' : serializer.data,
            'message' : 'Activity not added'
        })
    

def startWork(request):
    
    serializer = WorkSerializer(data=request.data)

    if serializer.is_valid():
        serializer.save()
        return Response({
            'status' : True,
            'data' : serializer.data,
            'message' : 'Work Started'
        })
    else:
        return Response({
            'status' : False,
            'data' : serializer.data,
            'message' : 'Work not Started'
        })
    

def checkWork(request):

    username = request.query_params.get('username')
    date = request.query_params.get('date')

    if Work.objects.filter(username=username, startDate = date).exists() is True:

        filteredData = Work.objects.get(username=username, startDate = date)
        serializer = WorkSerializer(instance = filteredData, many=False)

        if(serializer.data['endTime'] == ''):
            return Response({
                'data' : {
                    'started_work' : True,
                    'ended_work' : False,
                    'start_time' : serializer.data['startTime'],
                    'start_date' : serializer.data['startDate'],
                    'id' : serializer.data['id']
                },
                'status' : status.HTTP_201_CREATED,
                'message' : f"{username} is working"
            })
        else:
            return Response({
                'data' : {
                    'started_work' : True,
                    'ended_work' : True,
                    'start_time' : '',
                    'start_date' : '',
                    'id' : serializer.data['id']
                },
                'status' : status.HTTP_201_CREATED,
                'message' : f'{username} finished work'
            }) 
    else:
        return Response({
            'data' : {
                'started_work' : False,
                'ended_work' : False,
                'start_time' : '',
                'start_date' : '',
                'id' : -1
            },
            'status' : status.HTTP_201_CREATED,
            'message' : 'Work not started'
        })


def endWork(request, pk):

    data = request.data
    work = Work.objects.get(id=pk)
    serializer = WorkSerializer(instance=work, data=data)

    if serializer.is_valid():
        serializer.save()
        return Response({
            'status' : True,
            'data' : serializer.data,
            'message' : 'Work updated'
        })
    else:
        return Response({
            'status' : False,
            'data' : serializer.data,
            'message' : 'Work not updated'
        })
    

def updateUser(request, pk):

    users = AppUsers.objects.get(id=pk)
    serializer = UserSerializer(instance=users, data=request.data)

    if serializer.is_valid():
        serializer.save()
        return Response({
            'status' : True,
            'data' : serializer.data,
            'message' : 'User details updated'
        })
    else:
        return Response({
            'status' : False,
            'data' : serializer.data,
            'message' : 'User details not updated'
        })
    

def getUsersList(request):

    users = AppUsers.objects.all()
    serializer = UserSerializer(users, many=True)

    return Response({
        'data' : serializer.data,
        'status' : True,
        'message' : 'Fetched Users list'
    })


def addAmountEntry(request):
    
    serializer = AmountSerializer(data=request.data)

    if serializer.is_valid():
        serializer.save()
        return Response({
            'status' : True,
            'data' : serializer.data,
            'message' : 'New Entry added'
        })
    else:
        return Response({
            'status' : False,
            'data' : serializer.data,
            'message' : 'Amount not added'
        })
    

def getAmountEntries(request):

    amount = Amount.objects.all()
    serializer = AmountSerializer(amount, many=True)

    totalReceivedAmount = 0
    totalSpentAmount = 0
    totalRemainingAmount = 0
    totalReimbursementAmt = 0

    for data in serializer.data:
        totalReceivedAmount += int(0 if data['receivedAmount'] == '' else data['receivedAmount'])
        totalSpentAmount += int(0 if data['spentAmount'] == '' else data['spentAmount'])
        totalRemainingAmount += int(0 if data['remainingAmount'] == '' else data['remainingAmount'])
        totalReimbursementAmt += int(0 if data['reimbursementamt'] == '' else data['reimbursementamt'])

    return Response({
        'status' : True,
        'data' : {
            'total_data' : {
                'total_received_amount' : totalReceivedAmount,
                'total_spent_amount' : totalSpentAmount,
                'total_remaining_amount' : totalRemainingAmount,
                'total_reimbursement_amount' : totalReimbursementAmt
            },
            'data_list' : serializer.data
        },
        'message' : 'Amount details retrieved'
    })


def updateAmountEntry(request, pk):

    data = request.data
    amount = Amount.objects.get(id=pk)
    serializer = AmountSerializer(instance=amount, data=data)

    if serializer.is_valid():
        serializer.save()
        return Response({
            'status' : True,
            'data' : serializer.data,
            'message' : 'Amount Entry details updated'
        })
    else:
        return Response({
            'status' : False,
            'data' : serializer.data,
            'message' : 'Amount Entry details not updated'
        })
    

def updateActivity(request, pk):

    data = request.data
    activity = Posts.objects.get(id=pk)
    serializer = ActivitySerializer(instance=activity, data=data)

    if serializer.is_valid():
        serializer.save()
        return Response({
            'status' : True,
            'data' : serializer.data,
            'message' : 'Activity updated'
        })
    else:
        return Response({
            'status' : False,
            'data' : serializer.data,
            'message' : 'Activity not updated'
        })
    

def addFood(request):

    scheduled = request.data['scheduled']
    date = request.data['date']
    incharge = request.data['incharge']

    if scheduled.lower() == 'all':
        if Food.objects.filter(date=date).exists() is True:
            return Response({
                'data' : {},
                'status' : False,
                'message' : f'{date} already booked'
            })
        else:

            types = ['Morning', 'Afternoon', 'Night']

            for i in types:

                data = {
                    'name' : request.data['name'],
                    'sponserName' : request.data['sponserName'],
                    'mobile' : request.data['mobile'],
                    'preparation' : request.data['preparation'],
                    'amount' : request.data['amount'],
                    'status' : request.data['status'],
                    'paymentMethod' : request.data['paymentMethod'],
                    'scheduled' : i,
                    'date' : request.data['date'],
                    'incharge' : request.data['incharge']
                }

                serializer = FoodSerializer(data=data)

                if serializer.is_valid():
                    serializer.save()

            return Response({
                'status' : True,
                'data' : serializer.data,
                'message' : 'Food booked to provide'
            })
            
    else:
        if Food.objects.filter(scheduled=scheduled, date=date, incharge=incharge).exists() is True:
            return Response({
                'data' : {},
                'status' : False,
                'message' : f'{scheduled} already booked'
            })
        else:
            serializer = FoodSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'status' : True,
                    'data' : serializer.data,
                    'message' : 'Food booked to provide'
                })
            else:
                return Response({
                    'status' : False,
                    'data' : serializer.data,
                    'message' : 'Food not booked to provide'
                })
        

def getFoodData(request):

    food = Food.objects.all()
    serializer = FoodSerializer(food, many=True)

    return Response({
        'status' : True,
        'data' : serializer.data,
        'message' : 'Food details retrieved'
    })


def updateFood(request, pk):

    activity = Food.objects.get(id=pk)
    serializer = FoodSerializer(instance=activity, data=request.data)

    if serializer.is_valid():
        serializer.save()
        return Response({
            'status' : True,
            'data' : serializer.data,
            'message' : 'Food details updated'
        })
    else:
        return Response({
            'status' : False,
            'data' : serializer.data,
            'message' : 'Food details not updated'
        })
    

def addDonation(request):
    
    serializer = DonationSerializer(data=request.data)

    if serializer.is_valid():
        serializer.save()
        return Response({
            'status' : True,
            'data' : serializer.data,
            'message' : 'New Item Donated'
        })
    else:
        return Response({
            'status' : False,
            'data' : serializer.data,
            'message' : 'Item not Donated'
        })
    

def getDonations(request):

    donation = Donation.objects.all()
    serializer = DonationSerializer(donation, many=True)

    return Response({
        'data' : serializer.data,
        'status' : True,
        'message' : 'Donation list retrieved'
    })


def updateDonation(request, pk):

    activity = Donation.objects.get(id=pk)
    serializer = DonationSerializer(instance=activity, data=request.data)

    if serializer.is_valid():
        serializer.save()
        return Response({
            'status' : True,
            'data' : serializer.data,
            'message' : 'Donation details updated'
        })
    else:
        return Response({
            'status' : False,
            'data' : serializer.data,
            'message' : 'Donation details not updated'
        })
    

def getAttendance(request):

    type = request.query_params.get('type')

    if type == 'date':

        users = AppUsers.objects.all()
        userSerializer = UserSerializer(users, many=True)

        date = request.query_params.get('date')
        work = Work.objects.filter(startDate=date)
        serializer = WorkSerializer(instance = work, many= True)

        workData = serializer.data

        dailyAttendance = []
        present = 0
        absent = 0

        for data in userSerializer.data:
            if data['type'] != 'Director':
                if any(data['username'] == d['username'] for d in workData):
                    i = next((index for index, entry in enumerate(workData) if entry['username'] == data['username']), -1)
                    dailyAttendance.append({
                        'name' : data['username'],
                        'info' :{
                            'start_time' : workData[i]['startTime'],
                            'end_time' : workData[i]['endTime']
                        }
                    })
                    present += 1
                else:
                    dailyAttendance.append({
                        'name' : data['username'],
                        'info' : {
                            'start_time' : '-',
                            'end_time' : '-'
                        }
                    })
                    absent += 1
        return Response({
            'status' : True,
            'data' : {
                'attendance' : dailyAttendance,
                'attendance_info' : {
                    'present' : present,
                    'absent' : absent
                }
            },
            'message' : 'Attendance list retrived'
        })
    elif type == 'user':
      
        user = request.query_params.get('username')
      
        work = Work.objects.filter(username=user)
        serializer = WorkSerializer(instance = work, many= True)
      
        user = AppUsers.objects.get(username=user)
        userSerializer = UserSerializer(instance = user, many= False)
      
        return Response({
            'status' : True,
            'data' : {
                'user_data' : userSerializer.data,
                'work_data' : serializer.data,
            },
            'message' : 'Attendance list retrieved'
        })
    
    else:
        return Response({
            'status' : False,
            'data' : { },
            'message' : 'Attendance list retrived'
        })
    

def getProjectInfo(request):

    projects = Projects.objects.all()
    serializer = ProjectSerializer(projects, many=True)

    amount = Amount.objects.all()
    amountSerializer = AmountSerializer(amount, many=True)

    totalSents = 0
    projectSents = {}
    projectImages = {}

    D3 = {}

    for data in amountSerializer.data:
        totalSents += int(0 if data['spentAmount'] == '' else data['spentAmount'])
        projectSents[data['projectName']] = int(0 if data['spentAmount'] == '' else data['spentAmount'] 
        if data['projectName'] not in projectSents else projectSents[data['projectName']] + int(0 if data['spentAmount'] == '' else data['spentAmount']))

    for data in serializer.data:
        if data['name'] not in projectSents:
            projectSents[data['name']] = 0
        projectImages[data['name']] = data['image']

    projectDetails = [{'name': key, 'sents' : projectSents[key], 'image' : projectImages[key]} for key in projectSents.keys()]

    return Response({
        'status' : True,
        'data' : { 
            'project_count' : len(serializer.data),
            'total_sent' : totalSents,
            'project_details' : projectDetails
        },
        'message' : 'Project Details Retrieved'
    })


def getProjectIncharge(request):

    username = request.query_params.get('username')

    if Projects.objects.filter(coordinatorName=username).exists():

        projects = Projects.objects.get(coordinatorName = username)
        serializer = ProjectSerializer(projects, many=False)

        return Response({
            'status' : True,
            'data' : serializer.data,
            'message' : 'Project Details Retrieved'
        })

    else:
        return Response({
            'status' : True,
            'data' : {},
            'message' : 'Project Details Retrieved'
        })


def getProjectAmountEntries(request):

    projectName = request.query_params.get('project_name')

    amount = Amount.objects.filter(projectName = projectName)
    serializer = AmountSerializer(amount, many=True)

    totalReceivedAmount = 0
    totalSpentAmount = 0
    totalRemainingAmount = 0
    totalReimbursementAmt = 0

    for data in serializer.data:
        totalReceivedAmount += int(0 if data['receivedAmount'] == '' else data['receivedAmount'])
        totalSpentAmount += int(0 if data['spentAmount'] == '' else data['spentAmount'])
        totalRemainingAmount += int(0 if data['remainingAmount'] == '' else data['remainingAmount'])
        totalReimbursementAmt += int(0 if data['reimbursementamt'] == '' else data['reimbursementamt'])

    return Response({
        'status' : True,
        'data' : {
            'total_data' : {
                'total_received_amount' : totalReceivedAmount,
                'total_spent_amount' : totalSpentAmount,
                'total_remaining_amount' : totalRemainingAmount,
                'total_reimbursement_amount' : totalReimbursementAmt
            },
            'data_list' : serializer.data
        },
        'message' : 'Amount details retrieved'
    })


def getUsersListByProject(request):

    createdBy = request.query_params.get('created_user')

    users = AppUsers.objects.filter(created_by=createdBy)
    serializer = UserSerializer(users, many=True)

    activeUsers = AppUsers.objects.filter(created_by=createdBy, active = 1).count()

    return Response({
        'data' : {
            'total_data' : {
                'total_users' : len(serializer.data),
                'active_users' : activeUsers
            },
            'data_list' : serializer.data
        },
        'status' : True,
        'message' : 'Fetched Users list'
    })


def getDonationGraph(request):

    duration = request.query_params.get('duration')

    today = timezone.now().date()
    donation = Donation.objects.all()

    if duration == '6':
        sixMonthsAgo = today - timedelta(days=6*30)

        lastSixMonthsData = Donation.objects.filter(created__gte=sixMonthsAgo)

        serializer = DonationSerializer(lastSixMonthsData, many=True)

        labels = [data.month for data in donation]
        dct = defaultdict(int)

        for key in labels:
            dct[key] += 1

        d = dict(dct)

        donationGraph = [{'name': key, 'count' : d[key]} for key in d.keys()]

        return Response({
            'status' : True,
            'data' : {
                'list' : serializer.data,
                'graph_data' : donationGraph
            },
            'message' : 'Last Six Months data retrived'
        })
    
    elif duration == '12':

        oneYearAgo = today - timedelta(days=12*30)

        oneYearAgoData = Donation.objects.filter(created__gte=oneYearAgo)

        serializer = DonationSerializer(oneYearAgoData, many=True)

        labels = [data.month for data in donation]
        dct = defaultdict(int)

        for key in labels:
            dct[key] += 1

        d = dict(dct)

        donationGraph = [{'name': key, 'count' : d[key]} for key in d.keys()]

        return Response({
            'status' : True,
            'data' : {
                'list' : serializer.data,
                'graph_data' : donationGraph
            },
            'message' : 'Last 1 year data retrived'
        })
    
    elif duration == '60':
        
        fiveYeasrAgo = today - timedelta(days=60*30)
        
        fiveYearsAgoData = Donation.objects.filter(created__gte=fiveYeasrAgo)
        
        serializer = DonationSerializer(fiveYearsAgoData, many=True)
        
        labels = [data.month for data in donation]
        dct = defaultdict(int)

        for key in labels:
            dct[key] += 1

        d = dict(dct)

        donationGraph = [{'name': key, 'count' : d[key]} for key in d.keys()]

        return Response({
            'status' : True,
            'data' : {
                'list' : serializer.data,
                'graph_data' : donationGraph
            },
            'message' : 'Last 5 years data retrived'
        })
    
    else:
        
        labels = [data.month for data in donation]
        dct = defaultdict(int)

        return Response({
            'status' : True,
            'data' : {
                'list' : [],
                'graph_data' : []
            },
            'message' : 'No data found'
        })
    

def getFoodListByDate(request):

    date = request.query_params.get('date')

    food = Food.objects.filter(date=date)
    serializer = FoodSerializer(food, many=True)

    return Response({
        'data' : serializer.data,
        'status' : True,
        'message' : 'Fetched Food list'
    })
    

def getFoodListByDate(request):

    date = request.query_params.get('date')

    food = Food.objects.filter(date=date)
    serializer = FoodSerializer(food, many=True)

    return Response({
        'data' : serializer.data,
        'status' : True,
        'message' : 'Fetched Food list'
    })


def getProjectGraph(request):

    duration = request.query_params.get('duration')

    today = timezone.now().date()
    projects = Projects.objects.all()

    if duration == 'Last Month':

        oneMonthAgo = today - timedelta(days=30)
        
        lastMonthAmount = Amount.objects.filter(created__gte=oneMonthAgo)
        lastMonthData = Projects.objects.filter(created__gte=oneMonthAgo)

        amountSerializer = AmountSerializer(lastMonthAmount, many=True)
        serializer = ProjectSerializer(lastMonthData, many=True)

        totalSents = 0

        for data in amountSerializer.data:
            totalSents += int(0 if data['spentAmount'] == '' else data['spentAmount'])

        labels = [data.name for data in projects]
        dct = defaultdict(int)

        for key in labels:
            dct[key] += 1

        total_value = Amount.objects.aggregate(total=Sum('spentAmount'))['total']
        categories = Amount.objects.values('projectName').annotate(total_spent=Sum('spentAmount'))

        projectGraph = []

        for category in categories:

            percentage = (category['total_spent'] / total_value) * 100 if total_value else 0

            projectGraph.append({
                'name': category['projectName'],
                'percentage': round(percentage, 2)
            })

        return Response({
            'status' : True,
            'data' : {
                'total_sents' : totalSents,
                'data_list' : serializer.data,
                'graph_data' : projectGraph
            },
            'message' : 'Last Month data retrived'
        })
    
    elif duration == '6 Months':

        sixMonthsAgo = today - timedelta(days=6*30)
        
        lastSixMonthsAmount = Amount.objects.filter(created__gte=sixMonthsAgo)
        lastSixMonthsData = Projects.objects.filter(created__gte=sixMonthsAgo)

        amountSerializer = AmountSerializer(lastSixMonthsAmount, many=True)
        serializer = ProjectSerializer(lastSixMonthsData, many=True)

        totalSents = 0

        for data in amountSerializer.data:
            totalSents += int(0 if data['spentAmount'] == '' else data['spentAmount'])

        labels = [data.name for data in projects]
        dct = defaultdict(int)

        for key in labels:
            dct[key] += 1

        total_value = Amount.objects.aggregate(total=Sum('spentAmount'))['total']
        categories = Amount.objects.values('projectName').annotate(total_spent=Sum('spentAmount'))

        projectGraph = []

        for category in categories:

            percentage = (category['total_spent'] / total_value) * 100 if total_value else 0

            projectGraph.append({
                'name': category['projectName'],
                'percentage': round(percentage, 2)
            })

        return Response({
            'status' : True,
            'data' : {
                'total_sents' : totalSents,
                'data_list' : serializer.data,
                'graph_data' : projectGraph
            },
            'message' : 'Last Six months data retrived'
        })
    
    elif duration == 'All':
        
        amount = Amount.objects.all()
        projects = Projects.objects.all()

        amountSerializer = AmountSerializer(amount, many=True)
        serializer = ProjectSerializer(projects, many=True)

        totalSents = 0

        for data in amountSerializer.data:
            totalSents += int(0 if data['spentAmount'] == '' else data['spentAmount'])

        labels = [data.name for data in projects]
        dct = defaultdict(int)

        for key in labels:
            dct[key] += 1

        total_value = Amount.objects.aggregate(total=Sum('spentAmount'))['total']
        categories = Amount.objects.values('projectName').annotate(total_spent=Sum('spentAmount'))

        projectGraph = []

        for category in categories:

            percentage = (category['total_spent'] / total_value) * 100 if total_value else 0

            projectGraph.append({
                'name': category['projectName'],
                'percentage': round(percentage, 2)
            })

        return Response({
            'status' : True,
            'data' : {
                'total_sents' : totalSents,
                'data_list' : serializer.data,
                'graph_data' : projectGraph
            },
            'message' : 'All Data retrived'
        })
    
    else:
        return Response({
            'status' : True,
            'data' : {
                'total_sents' : 0,
                'data_list' : [],
                'graph_data' : {}
            },
            'message' : 'Data retrived'
        })
    

def generate_pdf(data, totals, project):

    fileName = f"amounts_{datetime.now().strftime("%d-%m-%Y")}.pdf"

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] =  f'attachment; filename={fileName}'
    
    buffer = SimpleDocTemplate(response, pagesize=letter)
   
    elements = []
    elements.append(Table([["PROJECT TITLE:", project]], colWidths=[200, 200]))
    elements.append(
        Table([[
            "Total Amount", f"{totals['received']:.2f}",
            "Spent Amount", f"{totals['spent']:.2f}",
            "Balance Amount", f"{totals['remaining']:.2f}",
            "Reimbursement Amount", f"{totals['reimbursement']:.2f}"
        ]])
    )

    headers = ['Date', 'Reason', 'Spending Amount']
    data_rows = [[amt.created.strftime('%d-%m-%Y'), amt.reason, f"{float(amt.spentAmount or 0.0):.2f}"] for amt in data]
    table_data = [headers] + data_rows
    table = Table(table_data)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    
    table.setStyle(style)
    elements.append(table)
    buffer.build(elements)
    
    return response


def generate_excel(data, totals, project):

    workbook = Workbook()
    sheet = workbook.active

    sheet['B1'] = 'PROJECT TITLE:'
    sheet['D1'] = project
    sheet.merge_cells('B1:C1')
    sheet.merge_cells('D1:E1')

    sheet['B3'] = 'Total Amount'
    sheet['C3'] = f"₹{totals['received']:.2f}"
    sheet['D3'] = 'Spent Amount'
    sheet['E3'] = f"₹{totals['spent']:.2f}"
    sheet['F3'] = 'Balance Amount'
    sheet['G3'] = f"₹{totals['remaining']:.2f}"
    sheet['H3'] = 'Reimbursement Amount'
    sheet['I3'] = f"₹{totals['reimbursement']:.2f}"

    headers = ['Date', 'Reason', 'Spending Amount']
    sheet.append(headers)

    for amt in data:

        row = [
            amt.created.strftime('%d-%m-%Y'),
            amt.reason,
            f"₹{float(amt.spentAmount or 0.0):.2f}"
        ]
        sheet.append(row)

    for column in sheet.columns:

        max_length = max(len(str(cell.value)) for cell in column if cell.value) + 2
        sheet.column_dimensions[get_column_letter(column[0].column)].width = max_length

    fileName = f"amounts{datetime.now().strftime("%d-%m-%Y")}.xlsx"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={fileName}'
    workbook.save(response)

    return response


def getProjectByProjectName(request):

    name = request.query_params.get('name')

    project = Projects.objects.get(name=name)
    serializer = ProjectSerializer(project, many=False)

    return Response({
        'data' : serializer.data,
        'status' : True,
        'message' : 'Fetched Project'
    })


def updateProject(request, pk):

    project = Projects.objects.get(id=pk)
    serializer = ProjectSerializer(instance=project, data=request.data)

    if serializer.is_valid():

        serializer.save()

        users = AppUsers.objects.get(username=request.data['coordinatorName'])
        userserializer = UserSerializer(users, many=False)

        data = {
            'username': userserializer.data['username'], 
            'number': userserializer.data['number'],
            'password': request.data['password'], 
            'token': userserializer.data['token'], 
            'type': userserializer.data['type'], 
            'active': userserializer.data['active'], 
            'created_by': userserializer.data['created_by'],
        }

        updateSerializer = UserSerializer(instance=users, data=data)

        if updateSerializer.is_valid():
            updateSerializer.save()
            return Response({
                'status' : True,
                'data' : updateSerializer.data,
                'message' : 'Projects details updated'
            })
        else:
            return Response({
                'status' : False,
                'data' : userserializer.data,
                'message' : 'Projects details not updated'
            })
        
    else:
        return Response({
            'status' : False,
            'data' : serializer.data,
            'message' : 'Projects details not updated'
        })